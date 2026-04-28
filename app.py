import streamlit as st
import openpyxl, re, unicodedata, os, io, base64
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pdfplumber

st.set_page_config(
    page_title="Corrector CO Natura",
    page_icon="📋",
    layout="centered"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&family=DM+Mono&display=swap');

* { font-family: 'DM Sans', sans-serif; }

.stApp {
    background: #0f1117;
    color: #e8e8e8;
}

h1 {
    font-size: 1.6rem !important;
    font-weight: 600 !important;
    color: #ffffff !important;
    letter-spacing: -0.5px;
}

.subtitle {
    color: #6b7280;
    font-size: 0.9rem;
    margin-top: -12px;
    margin-bottom: 32px;
}

.upload-card {
    background: #1a1d27;
    border: 1px solid #2a2d3a;
    border-radius: 12px;
    padding: 20px 24px;
    margin-bottom: 12px;
    transition: border-color 0.2s;
}

.upload-card:hover { border-color: #3d4155; }

.upload-label {
    font-size: 0.8rem;
    font-weight: 500;
    color: #9ca3af;
    text-transform: uppercase;
    letter-spacing: 1px;
    margin-bottom: 8px;
}

.file-ok {
    color: #34d399;
    font-size: 0.85rem;
    display: flex;
    align-items: center;
    gap: 6px;
}

section[data-testid="stFileUploadDropzone"] {
    background: #13151f !important;
    border: 1.5px dashed #2a2d3a !important;
    border-radius: 8px !important;
}

.stButton > button {
    background: #4f6ef7 !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    padding: 12px 32px !important;
    font-size: 0.95rem !important;
    font-weight: 500 !important;
    width: 100%;
    transition: all 0.2s !important;
}

.stButton > button:hover {
    background: #3d5ce0 !important;
    transform: translateY(-1px);
}

.stButton > button:disabled {
    background: #2a2d3a !important;
    color: #4b5563 !important;
}

.status-bar {
    background: #1a1d27;
    border: 1px solid #2a2d3a;
    border-radius: 8px;
    padding: 14px 18px;
    font-size: 0.85rem;
    color: #9ca3af;
    margin-top: 16px;
    font-family: 'DM Mono', monospace;
}

.divider {
    border: none;
    border-top: 1px solid #1e2130;
    margin: 24px 0;
}

.stDownloadButton > button {
    background: #1a2a1a !important;
    color: #34d399 !important;
    border: 1.5px solid #34d399 !important;
    border-radius: 8px !important;
    width: 100%;
    font-weight: 500 !important;
}
</style>
""", unsafe_allow_html=True)


# ── helpers ──────────────────────────────────────────────────────────────────

def norm(s):
    if s is None: return ""
    s = str(s).strip()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = re.sub(r'[/\.\-_,]', '', s)
    return re.sub(r'\s+', ' ', s).upper().strip()

def compare(a, b):
    return "✅ OK" if norm(str(a)) == norm(str(b)) else "❌ DIFERENCIA"

def compare_num(a, b):
    try:
        return "✅ OK" if abs(float(str(a).replace(',','.')) - float(str(b).replace(',','.'))) < 0.05 else "❌ DIFERENCIA"
    except:
        return "❌ ERROR"

def parse_num(s):
    s = str(s).strip()
    if re.search(r'\d\.\d{3}', s):
        s = s.replace('.', '').replace(',', '.')
    else:
        s = s.replace(',', '.')
    try: return float(s)
    except: return 0.0

# ── Groq Vision ──────────────────────────────────────────────────────────────

def pdf_a_imagenes_b64(path, dpi=200):
    """Convierte PDF a lista de imágenes en base64."""
    try:
        from pdf2image import convert_from_path
        pages = convert_from_path(path, dpi=dpi)
        imgs = []
        for page in pages:
            buf = io.BytesIO()
            page.save(buf, format='PNG')
            buf.seek(0)
            imgs.append(base64.standard_b64encode(buf.read()).decode('utf-8'))
        return imgs
    except Exception as e:
        return []

def groq_vision_co(path):
    """Usa Groq Vision para extraer texto del CO cuando pdfplumber falla."""
    try:
        from groq import Groq
        api_key = st.secrets.get("GROQ_API_KEY", "")
        if not api_key:
            return []
        client = Groq(api_key=api_key)
        imagenes = pdf_a_imagenes_b64(path)
        if not imagenes:
            return []

        texto_completo = []
        for i, img_b64 in enumerate(imagenes):
            response = client.chat.completions.create(
                model="meta-llama/llama-4-scout-17b-16e-instruct",
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "image_url",
                                "image_url": {"url": f"data:image/png;base64,{img_b64}"}
                            },
                            {
                                "type": "text",
                                "text": (
                                    "Este es un Certificado de Origen (CO) de comercio exterior. "
                                    "Extraé TODO el texto visible en la imagen, manteniendo la estructura línea por línea. "
                                    "Es muy importante preservar: números de ítem, códigos NCM (formato XXXX.XX.XX), "
                                    "cantidades con formato numérico (ej: 1.234,567), unidades (gr, kg, pc), "
                                    "valores monetarios, y códigos de material (números de 7-8 dígitos después de ';'). "
                                    "No agregues interpretación, solo el texto exacto línea por línea."
                                )
                            }
                        ]
                    }
                ],
                max_tokens=4096
            )
            texto_completo.extend(response.choices[0].message.content.split('\n'))

        return texto_completo
    except Exception:
        return []

def groq_vision_fc(path):
    """Usa Groq Vision para extraer fecha y total EXW de la FC cuando pdfplumber falla."""
    try:
        from groq import Groq
        api_key = st.secrets.get("GROQ_API_KEY", "")
        if not api_key:
            return None
        client = Groq(api_key=api_key)
        imagenes = pdf_a_imagenes_b64(path)
        if not imagenes:
            return None

        # Solo procesamos la primera página para FC (suele ser suficiente)
        img_b64 = imagenes[0]
        response = client.chat.completions.create(
            model="meta-llama/llama-4-scout-17b-16e-instruct",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image_url",
                            "image_url": {"url": f"data:image/png;base64,{img_b64}"}
                        },
                        {
                            "type": "text",
                            "text": (
                                "Esta es una Factura Comercial (FC) de comercio exterior. "
                                "Extraé TODO el texto visible línea por línea, preservando exactamente: "
                                "fechas (formato DD/MM/YYYY), totales EXW con sus valores numéricos, "
                                "totales ARS, y cualquier número monetario. "
                                "No agregues interpretación, solo el texto exacto línea por línea."
                            )
                        }
                    ]
                }
            ],
            max_tokens=2048
        )
        return response.choices[0].message.content.split('\n')
    except Exception:
        return None

# ── texto suficiente? ────────────────────────────────────────────────────────

def texto_es_suficiente(lines, min_chars=100):
    """Determina si el texto extraído por pdfplumber es suficiente."""
    total = sum(len(l.strip()) for l in lines)
    return total >= min_chars

# ── leer archivos ────────────────────────────────────────────────────────────

def leer_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws_item = wb['Item']
    headers = [ws_item.cell(1, c).value for c in range(1, 50)]
    col_cant = next((i+1 for i,h in enumerate(headers) if h and 'CANTIDAD' in str(h).upper()), 7)
    col_mat  = next((i+1 for i,h in enumerate(headers) if h and 'MARCA-MODEL' in str(h).upper()), 6)
    items = []
    for row in ws_item.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            ncm_raw = str(row[1]) if row[1] else ''
            ncm_clean = ncm_raw[:10]
            mat = row[col_mat-1]
            if mat and not str(mat).replace('.','').isdigit():
                mat = row[5]
            items.append({'ITEM': row[0], 'NCM': ncm_clean, 'CANTIDAD': row[col_cant-1], 'MARCA_MODEL_OTRO': mat})
    ws_car = wb['Carátula']
    rows_car = list(ws_car.iter_rows(values_only=True))
    empresa  = rows_car[1][2] if len(rows_car) > 1 else None
    facturas = rows_car[4][0] if len(rows_car) > 4 else None
    vendedor = rows_car[4][1] if len(rows_car) > 4 else None
    return {'items': items, 'empresa': empresa, 'facturas': facturas, 'vendedor': vendedor}

def leer_fc(path):
    # Intento 1: pdfplumber
    full = ''
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t: full += t + '\n'

    lines = full.split('\n')

    # Fallback Groq si el texto es insuficiente
    if not texto_es_suficiente(lines):
        groq_lines = groq_vision_fc(path)
        if groq_lines:
            lines = groq_lines

    data = {'fecha': '', 'total_exw': None, 'total_ars': False}
    for l in lines:
        m = re.search(r'FECHA\s+(\d{2}/\d{2}/\d{4})', l, re.IGNORECASE)
        if m: data['fecha'] = m.group(1)
        m = re.search(r'TOTAL\s+EXW\s+([\d\.,]+)', l, re.IGNORECASE)
        if m: data['total_exw'] = parse_num(m.group(1))
        if re.search(r'TOTAL\s+ARS', l, re.IGNORECASE): data['total_ars'] = True
    return data

def leer_co_pdf(path):
    # Intento 1: pdfplumber
    full_lines = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t: full_lines.extend(t.split('\n'))

    # Fallback Groq si el texto es insuficiente
    if not texto_es_suficiente(full_lines):
        groq_lines = groq_vision_co(path)
        if groq_lines:
            full_lines = groq_lines

    produtor = ''
    for l in full_lines[:40]:
        if re.search(r'(industria|comercio|ltda)', l, re.IGNORECASE):
            if not re.search(r'(certificado|origen|validez)', l, re.IGNORECASE):
                produtor = re.split(r'(RODOVIA|ROD\.|RUA|AV\.|\bSP\b)', l, flags=re.IGNORECASE)[0].strip()
                if produtor: break

    importador = ''
    for i, l in enumerate(full_lines):
        if re.search(r'2[,\.]\s*importador', l, re.IGNORECASE):
            for j in range(i+1, min(i+6, len(full_lines))):
                c = full_lines[j].strip()
                if c and any(k in c.upper() for k in ['NATURA', 'S/A', 'LTDA', 'SA']):
                    importador = re.split(r'(CAZADORES|AV\.|RUA|\d{5})', c, flags=re.IGNORECASE)[0].strip()
                    break
            break

    factura_num = data_co = ''
    for l in full_lines:
        ln = unicodedata.normalize('NFD', l)
        ln = ''.join(c for c in ln if unicodedata.category(c) != 'Mn')
        m = re.search(r'[Nn]um[^\s:]*[:\s]+([A-Z]{0,5}\d{6,12})', ln)
        if m and not factura_num: factura_num = m.group(1)
        m = re.search(r'[Dd]ata[:\s]+(\d{2}/\d{2}/\d{4})', l)
        if m and not data_co: data_co = m.group(1)

    pattern = re.compile(
        r'^\s*(\d{1,2})\s+(\d{4}\.\d{2}\.\d{2})[^\n]*?([\d\.]+,\d{3})\s+(?:gr|kg|pc|p[çc°¢])\s+([\d\.]+,\d{3})'
    )

    mat_re_semicolon = re.compile(r';\s*(\d{7,8})(?:\s|$)')   # con ";" como pista fuerte
    mat_re_solosemi  = re.compile(r'^\s*;\s*(\d{7,8})\s*$')   # línea que empieza con ";"
    mat_re_candidato = re.compile(r'(?<!\d)(\d{7,8})(?!\d)')  # fallback: 7-8 dígitos exactos
    mat_re_ncm       = re.compile(r'\d{4}\.\d{2}\.\d{2}')
    mat_re_djo       = re.compile(r'\d{6,8}\s*[-•]\s*\d{2}/\d{2}/\d{4}')

    def es_material_valido(linea):
        if mat_re_ncm.search(linea): return False
        if mat_re_djo.search(linea): return False
        if re.search(r'\d+,\d+', linea): return False
        return True

    def buscar_material(lines, start, window=50):
        end = min(start + window, len(lines))

        # Paso 1: buscar con ";" como pista fuerte
        for j in range(start, end):
            mm = mat_re_semicolon.search(lines[j])
            if mm: return int(mm.group(1))
            mm = mat_re_solosemi.match(lines[j])
            if mm: return int(mm.group(1))

        # Paso 2: fallback sin ";", descartando NCM/DJO/cantidades
        for j in range(start, end):
            linea = lines[j]
            if not es_material_valido(linea): continue
            mm = mat_re_candidato.search(linea)
            if mm: return int(mm.group(1))

        return None

    items = []
    for i, l in enumerate(full_lines):
        m = pattern.match(l)
        if m:
            orden, ncm, cant_str, val_str = int(m.group(1)), m.group(2), m.group(3), m.group(4)
            material = buscar_material(full_lines, i)
            if not any(it['orden'] == orden for it in items):
                items.append({'orden': orden, 'ncm': ncm, 'cantidad': cant_str,
                              'cantidad_num': parse_num(cant_str), 'valor': parse_num(val_str),
                              'material': material})

    materiales_encontrados = {it['material'] for it in items if it['material']}
    for i, l in enumerate(full_lines):
        mm = mat_re_inline.search(l)
        if not mm and i > 0 and ';' in full_lines[i-1]:
            mm = mat_re_nextline.match(l)
        if mm:
            mat = int(mm.group(1))
            if mat not in materiales_encontrados:
                for back in range(i-1, max(i-60, -1), -1):
                    m = re.search(r'(\d{4}\.\d{2}\.\d{2})[^\n]*?([\d\.]+,\d{3})\s+(?:gr|kg|pc|p[çc°¢])\s+([\d\.]+,\d{3})', full_lines[back])
                    if m:
                        ncm, cant_str = m.group(1), m.group(2)
                        for it in items:
                            if it['ncm'] == ncm and it['cantidad'] == cant_str and it['material'] is None:
                                it['material'] = mat
                                materiales_encontrados.add(mat)
                                break
                        break

    obs_lines = []
    capture = False
    for l in full_lines:
        if re.search(r'12\.\s*observa', l, re.IGNORECASE): capture = True; continue
        if capture:
            if re.search(r'(certificac|declarac|13\.|14\.)', l, re.IGNORECASE): break
            if l.strip(): obs_lines.append(l.strip())
    obs = ' '.join(obs_lines).strip()

    return {'produtor': produtor, 'importador': importador, 'factura_num': factura_num,
            'data': data_co, 'items': items, 'observaciones': obs}


def generar_reporte(xl, fc_data, co, op_id):
    hdr_font   = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    hdr_fill   = PatternFill('solid', start_color='1F4E79')
    ok_fill    = PatternFill('solid', start_color='C6EFCE')
    err_fill   = PatternFill('solid', start_color='FFC7CE')
    warn_fill  = PatternFill('solid', start_color='FFEB9C')
    na_fill    = PatternFill('solid', start_color='EDEDED')
    normal_font = Font(name='Arial', size=10)
    bold_font   = Font(name='Arial', bold=True, size=10)
    thin   = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    def style_hdr(c):    c.font=hdr_font; c.fill=hdr_fill; c.alignment=center; c.border=border
    def style_normal(c): c.font=normal_font; c.alignment=left; c.border=border
    def style_result(c, v):
        c.font=bold_font; c.alignment=center; c.border=border
        if '✅' in v:   c.fill=ok_fill
        elif '❌' in v: c.fill=err_fill
        elif '⚠️' in v: c.fill=warn_fill
        else:            c.fill=na_fill
    def style_section(ws, row, text):
        ws.merge_cells(f'A{row}:G{row}')
        c=ws[f'A{row}']; c.value=text
        c.font=Font(name='Arial', bold=True, size=11, color='FFFFFF')
        c.fill=PatternFill('solid', start_color='2E75B6')
        c.alignment=center; c.border=border; ws.row_dimensions[row].height=22
    def write_row(ws, row, data, result_col):
        for col, val in enumerate(data, start=1):
            c = ws.cell(row=row, column=col, value=val)
            if col == result_col: style_result(c, str(val))
            else: style_normal(c)
        ws.row_dimensions[row].height = 20

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Reporte Cruces'
    ws.merge_cells('A1:G1')
    ws['A1'] = f'REPORTE DE VALIDACIÓN — OPERACIÓN {op_id}'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', start_color='1F3864')
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 30
    for col, w in zip('ABCDEFG', [18, 28, 32, 32, 16, 35, 10]):
        ws.column_dimensions[col].width = w

    from collections import defaultdict
    co_by_material = defaultdict(list)
    for ci in co['items']:
        if ci['material']:
            co_by_material[ci['material']].append(ci)

    def buscar_co_item(mat_int, cant_excel):
        candidatos = co_by_material.get(mat_int, [])
        if not candidatos:
            return None
        for ci in candidatos:
            cq = ci['cantidad_num']
            if abs(cant_excel - cq) < 0.01: return ci
            if abs(cant_excel * 1000 - cq) < 0.5: return ci
            if abs(cant_excel / 1000 - cq) < 0.0001: return ci
        return candidatos[0]

    row = 3
    style_section(ws, row, 'LÓGICA 1 — Excel (Solapa Item) vs PDF CO'); row += 1
    for col, h in enumerate(['ITEM / Material','CAMPO','VALOR EXCEL','VALOR PDF CO','RESULTADO','CONSIDERACIÓN',''], start=1):
        style_hdr(ws.cell(row=row, column=col, value=h))
    row += 1
    for item in xl['items']:
        ref = item['MARCA_MODEL_OTRO']
        cant_exc = parse_num(str(item['CANTIDAD']))
        co_item = buscar_co_item(int(ref), cant_exc) if ref and str(ref).replace('.','').isdigit() else None
        ncm_display = str(item['NCM'])[:10] if item['NCM'] else ''
        ncm_10 = ncm_display.replace('.','')
        if co_item:
            res_ncm  = "✅ OK" if ncm_10 == co_item['ncm'].replace('.','') else "❌ DIFERENCIA"
            cq = co_item['cantidad_num']
            match_cant = (abs(cant_exc - cq) < 0.01 or
                          abs(cant_exc * 1000 - cq) < 0.5 or
                          abs(cant_exc / 1000 - cq) < 0.0001)
            res_cant = "✅ OK" if match_cant else "❌ DIFERENCIA"
            cant_excel_str = str(cant_exc) if cant_exc != int(cant_exc) else str(int(cant_exc))
            rows_data = [
                (f"{item['ITEM']} / {ref}", 'NCM', ncm_display, co_item['ncm'], res_ncm, '10 primeros caracteres'),
                ('', 'CANTIDAD', cant_excel_str, co_item['cantidad'], res_cant, 'Tolera conversión kg↔gr'),
            ]
        else:
            rows_data = [
                (f"{item['ITEM']} / {ref}", 'NCM', ncm_display, '⚠️ No encontrado en CO', '⚠️ SIN MATCH', 'Material no hallado'),
                ('', 'CANTIDAD', str(item['CANTIDAD']), '⚠️ No encontrado en CO', '⚠️ SIN MATCH', ''),
            ]
        for d in rows_data: write_row(ws, row, d, 5); row += 1

    row += 1
    style_section(ws, row, 'LÓGICA 2 — Excel (Solapa Carátula) vs PDF CO'); row += 1
    for col, h in enumerate(['CAMPO EXCEL','VALOR EXCEL','CAMPO PDF CO','VALOR PDF CO','RESULTADO','CONSIDERACIÓN',''], start=1):
        style_hdr(ws.cell(row=row, column=col, value=h))
    row += 1
    l2 = [
        ('FACTURAS', str(xl['facturas']), 'FACTURA COMERCIAL (Nro)', co['factura_num'], compare(str(xl['facturas']), co['factura_num']), 'Número de factura'),
        ('EMPRESA',  str(xl['empresa']),  '2. IMPORTADOR (razón social)', co['importador'], compare(str(xl['empresa']), co['importador']), 'Solo razón social'),
    ]
    if xl['vendedor']:
        l2.append(('VENDEDOR', str(xl['vendedor']), '1. PRODUTOR FINAL OU EXPORTADOR', co['produtor'], compare(str(xl['vendedor']), co['produtor']), 'Solo razón social'))
    else:
        l2.append(('VENDEDOR', 'NO INFORMADO EN EXCEL', '1. PRODUTOR FINAL OU EXPORTADOR', co['produtor'], '⚠️ NO APLICA', 'Campo ausente en este Excel'))
    for d in l2: write_row(ws, row, d, 5); ws.row_dimensions[row].height=22; row += 1

    row += 1
    style_section(ws, row, 'LÓGICA 3 — PDF CO vs PDF FC'); row += 1
    for col, h in enumerate(['CAMPO PDF CO','VALOR PDF CO','CAMPO PDF FC','VALOR PDF FC','RESULTADO','CONSIDERACIÓN',''], start=1):
        style_hdr(ws.cell(row=row, column=col, value=h))
    row += 1
    suma_co = sum(ci['valor'] for ci in co['items'])
    res_total = compare_num(suma_co, fc_data['total_exw']) if fc_data['total_exw'] else '⚠️ NO ENCONTRADO'
    l3 = [
        ('DATA', co['data'], 'FECHA', fc_data['fecha'], compare(co['data'], fc_data['fecha']), 'Sin consideraciones'),
        ('SUMA TOTAL "9. VALOR"', f"{suma_co:,.3f}", 'TOTAL EXW', f"{fc_data['total_exw']:,.2f}" if fc_data['total_exw'] else 'N/A', res_total, 'Suma todos los valores punto 9'),
    ]
    for d in l3: write_row(ws, row, d, 5); ws.row_dimensions[row].height=22; row += 1

    row += 1
    style_section(ws, row, 'LÓGICA 4 — PDF FC vs PDF CO (Campo 12. Observações)'); row += 1
    for col, h in enumerate(['CONDICIÓN (PDF FC)','ESTADO','LEYENDA ESPERADA EN CO-12','VALOR ENCONTRADO EN CO-12','RESULTADO','CONSIDERACIÓN',''], start=1):
        style_hdr(ws.cell(row=row, column=col, value=h))
    row += 1
    obs = co['observaciones']
    c12 = obs[:100] if obs else '⚠️ Campo 12 no encontrado en PDF'
    nota = 'Campo 12 leído correctamente' if obs else '⚠️ Subir CO con texto seleccionable'
    res_ars = ('✅ OK' if 'PESOS' in norm(obs) or 'REAIS' in norm(obs) else '❌ DIFERENCIA') if obs else '⚠️ NO VERIFICABLE'
    res_exw = ('✅ OK' if 'EXW' in norm(obs) else '❌ DIFERENCIA') if obs else '⚠️ NO VERIFICABLE'
    if fc_data['total_ars']:
        write_row(ws, row, ['FC tiene TOTAL ARS', 'Sí - por ítem', 'AO VALOR EM MOEDA LOCAL (PESOS OU REAIS)', c12, res_ars, nota], 5)
        ws.row_dimensions[row].height=30; row += 1
    if fc_data['total_exw']:
        write_row(ws, row, ['FC tiene TOTAL EXW', f"{fc_data['total_exw']:,.2f}", 'AO VALOR EXW DA FATURA COMERCIAL', c12, res_exw, nota], 5)
        ws.row_dimensions[row].height=30; row += 1

    row += 2
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = 'Nota: Comparaciones flexibles (ignoran acentos, espacios, mayúsculas, /, ., -)'
    ws[f'A{row}'].font = Font(name='Arial', italic=True, size=9, color='595959')
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── UI ───────────────────────────────────────────────────────────────────────

st.markdown("# Corrector CO Natura")
st.markdown('<p class="subtitle">Cruce automático Excel · CO · FC</p>', unsafe_allow_html=True)
st.markdown('<hr class="divider">', unsafe_allow_html=True)

col1, col2, col3 = st.columns(3)

with col1:
    st.markdown('<div class="upload-label">📊 Excel</div>', unsafe_allow_html=True)
    excel_file = st.file_uploader("Excel", type=["xlsx","xls"], label_visibility="collapsed", key="excel")
    if excel_file:
        st.markdown(f'<div class="file-ok">✓ {excel_file.name}</div>', unsafe_allow_html=True)

with col2:
    st.markdown('<div class="upload-label">📄 CO</div>', unsafe_allow_html=True)
    co_file = st.file_uploader("CO", type=["pdf"], label_visibility="collapsed", key="co")
    if co_file:
        st.markdown(f'<div class="file-ok">✓ {co_file.name}</div>', unsafe_allow_html=True)

with col3:
    st.markdown('<div class="upload-label">📄 FC</div>', unsafe_allow_html=True)
    fc_file = st.file_uploader("FC", type=["pdf"], label_visibility="collapsed", key="fc")
    if fc_file:
        st.markdown(f'<div class="file-ok">✓ {fc_file.name}</div>', unsafe_allow_html=True)

st.markdown('<hr class="divider">', unsafe_allow_html=True)

todos_ok = excel_file and co_file and fc_file

if todos_ok:
    if st.button("⚡ Generar Reporte"):
        with st.spinner("Procesando documentos..."):
            try:
                import tempfile
                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
                    f.write(excel_file.read()); excel_path = f.name
                with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as f:
                    f.write(co_file.read()); co_path = f.name
                with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as f:
                    f.write(fc_file.read()); fc_path = f.name

                st.markdown('<div class="status-bar">📊 Leyendo Excel...</div>', unsafe_allow_html=True)
                xl = leer_excel(excel_path)

                st.markdown('<div class="status-bar">📄 Leyendo FC...</div>', unsafe_allow_html=True)
                fc_data = leer_fc(fc_path)

                st.markdown('<div class="status-bar">📄 Procesando CO...</div>', unsafe_allow_html=True)
                co = leer_co_pdf(co_path)

                op_id = re.search(r'(\d{5,})', excel_file.name)
                op_id = op_id.group(1) if op_id else 'XXXX'

                buf = generar_reporte(xl, fc_data, co, op_id)

                for p in [excel_path, co_path, fc_path]:
                    try: os.unlink(p)
                    except: pass

                st.success(f"✅ Reporte generado — {len(co['items'])} ítems procesados")
                st.download_button(
                    label=f"⬇️ Descargar Reporte_{op_id}.xlsx",
                    data=buf,
                    file_name=f"Reporte_Cruces_{op_id}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"Error: {e}")
else:
    st.button("⚡ Generar Reporte", disabled=True)
    if not todos_ok:
        faltantes = []
        if not excel_file: faltantes.append("Excel")
        if not co_file: faltantes.append("CO")
        if not fc_file: faltantes.append("FC")
        st.markdown(f'<div class="status-bar">⏳ Faltan: {", ".join(faltantes)}</div>', unsafe_allow_html=True)
